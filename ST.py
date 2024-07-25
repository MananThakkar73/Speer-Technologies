import re
import requests
from openpyxl.reader.excel import load_workbook

global url_list
url_list = []
#method to get status code
def link_statusCode(wiki_url):
    return requests.get(wiki_url).status_code

#method to Accepts a Wikipedia link - return/throw an error if the link is not a valid wiki link
def wikilink(wiki_url):
    wiki_format = r'^(https?://)?(www\.)?([a-z]+)\.wikipedia\.org/wiki/([^\s]+)$'
    if re.match(wiki_format, wiki_url) and link_statusCode(wiki_url) == 200:
        #print("This is valid Wikipedia url")
        return True
    else:
        return Exception("Not valid Wikipedia Link")

#method to Accepts a valid integer between 1 to 3 - call it n
def acceptvalidint(n):
    if 1<= n <= 3:
        return True
    else:
        print("Please enter valid integer between 1 to 3")

#method to take wikipedia url from excel sheet
def datafromexcel(filename, sheet):
    datalist = []
    wb = load_workbook(filename=filename)
    sh = wb[sheet]

    row_count = sh.max_row
    col_count = sh.max_column
    row = []
    for r in range(2, row_count + 1):
        #row = []
        for j in range(1, col_count+1):
            datalist.append(sh.cell(row=r, column=j).value)
        #datalist.append(row)

    url = str(datalist).replace("'","").replace("[","").replace("]","").split(", ")

    return url

urls = datafromexcel("Wiki url.xlsx","Sheet1")

#Scrape the link provided in Step 1, for the first 10 unique (not previously added already) wiki links
#embedded in the excel and store them in a data structure of your choice.
def addToList(n):
    if acceptvalidint(n):

        for j in range(n):
            #for j in range(n):
            c=0
            for url in urls:
                if wikilink(url):
                    if url not in url_list:
                        url_list.append(url)
                        c+=1
                        if c == 10:
                            break

        print(url_list)
        print(len(url_list))

addToList(2)
